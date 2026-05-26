"""
세출예산 성질별 분류 자동화 프로그램 v5 (2026년 기준)
=========================================================
[변경사항]
- 2026년 통계목 체계 전면 반영 (PDF 별표11 기준)
- 자체/보조 구분: 국고+균특+기금+광역보조금 합계=0 → 자체사업
- 도시기본기능유지: 401-01 중 URBAN_INFRA_NAMES 산출근거명 자동 매칭
- 지방채 원리금상환 신규 대분류 추가 (311-xx, 601-05)
- 205-xx 전체 → 행정운영경비 > 의회비
- 행사실비지원금(301-11) → 경직성 이전경비
- 민간인국외여비(301-08) + 외빈초청여비(301-09) → 동일 항목
- 통계목 과목번호 이동 전면 반영
"""

# ─────────────────────────────────────────────────────────────────────────────
# 도시기본기능유지 기준 목록
# 2026년 예산서 401-01 중 노란색 셀 산출근거명 기준 (⊙/ｏ 기호 제거·공백 정규화)
# 다른 연도 파일 입력 시 이 목록으로 자동 매칭
# ─────────────────────────────────────────────────────────────────────────────
URBAN_INFRA_NAMES = {
    "2040년 의정부 도시기본계획 수립관련 검증수수료",
    "CCTV통합관제센터 및 방범용CCTV 유지관리",
    "가로등 권역별 유지보수",
    "가로수 관리 연간단가",
    "가로수 안전진단 및 처방",
    "가로수길 보행환경 정비사업",
    "가로화분 및 화단 식재사업",
    "건물번호판 설치",
    "걷고 싶은 보행로 확보 위한 동절기 장비 임차 용역",
    "공설묘지 유지보수",
    "공원 내 모래놀이터 소독사업",
    "공원 내 시설 유지보수 등 공사",
    "공원 등 수목 및 잔디 방제사업(1구역)",
    "공원 등 수목 및 잔디 방제사업(2구역)",
    "공원 등 운동기구 교체 공사(1구역)",
    "공원 등 운동기구 교체 공사(2구역)",
    "공원 등 잔디깎기 공사(1구역)",
    "공원 등 잔디깎기 공사(2구역)",
    "공원등 유지관리 보수(연간단가)",
    "공유재산 유지보수비",
    "광역교통정보시스템 유지관리",
    "교통사고 잔해물 처리 위탁용역(연간단가)",
    "교통신호 관제센터 유지관리 용역",
    "교통신호기 신설",
    "교통신호시설 유지보수(중랑천 동측 및 서측)",
    "교통안전표지 유지관리",
    "노후 교통신호제어기 교체",
    "노후 신호등 교체",
    "노후 철주 및 선로 교체",
    "녹지대 맨발길 등 유지관리 사업",
    "녹지대 시설물 정비사업",
    "녹지대 잔디깎기 및 관목전정",
    "녹지대 폐기물 처리 용역",
    "도로명판 설치",
    "도로사용료 감정평가 수수료",
    "도로안전시설물 설치 및 유지보수 (차선분리대 등)",
    "도시재생 계획수립 용역비",
    "동오 황토 맨발로(路) 유지관리",
    "동의교 외 1개소 보수보강공사",
    "동청사(주민센터) 시설관리",
    "디자인사업 시설물 유지 및 보수",
    "무인교통단속장비 교체",
    "문화시설 유지보수",
    "미지급 용지 보상금",
    "방범CCTV 이설 및 정비 공사",
    "방범CCTV 자가통신망 노후 광전송장비 교체",
    "배수펌프장 안점점검 용역(정밀·정기)",
    "배수펌프장, 배수문 시설물 유지보수",
    "배수펌프장, 저류지 시설물 유지관리(예초, 준설 등)",
    "버스정류장 시설물 유지관리",
    "버스정보시스템 유지관리",
    "보건소 청사 건축·토목·기계·전기·통신·소방 시설 유지관리",
    "보안등 권역별 유지보수",
    "보안등 신설 사업",
    "보행안전시설 유지관리",
    "비우로 도로정비공사",
    "사물주소판 설치",
    "산장연립 부지옹벽 계측관리 용역",
    "생태하천 가로등 유지보수",
    "생태하천 녹지대 유지관리",
    "생태하천 수해 폐기물 처리용역",
    "생태하천 안내판 정비사업(수해 복구)",
    "생태하천 조명시설 유지보수",
    "소규모 시설물 정기안전점검 용역",
    "소풍길 유지관리 사업",
    "소하천 및 구거 등 환경정비",
    "소하천 및 구거 예초 용역",
    "소하천 및 구거 유지보수(연간단가)",
    "소하천 소규모 준설 사업",
    "스마트도시 시스템 통합 유지관리",
    "승강기 유지보수",
    "시 일원 도로 제초 및 전지작업(단가계약)",
    "시 일원 도로구조물 정비공사(연간단가)",
    "시 일원 도로안전시설물 유지보수 공사 (연간단가)",
    "시 일원 도로표지판 정비 사업",
    "시 일원 보도 유지보수(연간단가)",
    "시 일원 빗물받이 유지보수(연간단가)",
    "시 일원 빗물받이 준설공사(연간단가)",
    "시 일원 자전거도로 유지보수(연간단가)",
    "시 일원 차도 유지보수공사(연간단가)",
    "시각장애인 음향신호기 설치 및 교체",
    "시일원 차선 유지보수",
    "신의교 빗물펌프장 개선공사",
    "어린이놀이시설 교체 공사",
    "의정부 소하천정비 종합계획(변경)수립 및 지형도면고시 용역",
    "의정부시 도로건설·관리계획(변경) 수립 용역",
    "의정부시 지역교통안전기본계획 수립용역",
    "의정부역 지하도상가 건축물 정밀안전진단 용역",
    "자가정보통신망 정비",
    "자전거 이용 활성화 계획 수립 용역",
    "자전거 이용시설 개선 사업",
    "제1·2종 시설물 안전점검 용역",
    "제3종 시설물 안전점검 용역",
    "종합운동장 주경기장 정밀안전점검",
    "중랑천 벽천분수 수경시설 보수",
    "지방하천 유지보수(연간단가)",
    "지방하천 준설공사(시비추가)",
    "지역현장 로드체킹 관리",
    "지진가속도 계측기 유지관리 용역",
    "지진가속도 계측기 유지보수",
    "지하보도 및 승강기 청소 용역",
    "지하보차도 및 빗물펌프장 시설물(기계, 전기,통신) 유지보수",
    "지하보차도 및 빗물펌프장 준설공사",
    "지하차도 및 빗물펌프장 수중펌프 교체",
    "천보로(금오동 420-114)일대 인도 콘크리트 정비",
    "청사유지보수",
    "체육시설 유지 및 보수",
    "택시정류장 환경 개선",
    "평화로 보도정비공사",
    "폭염그늘막 유지관리 용역",
    "하천 내 편의, 체육시설 등 신설 및 유지 보수",
    "하천 재해복구 및 유지관리 장비임차",
    "하천 제설 및 유지관리 장비임차",
    "행정정보통신망(구내통신선로) 구축",
    "현충탑 및 김풍익전투기념비 유지관리",
    "현충탑 및 김풍익전투기념비 조경관리",
    "현황측량 등",
    "호원천 생태하천 복원사업 사후모니터링 용역(1차년도)",
    "회룡역사거리 빗물펌프장 비상발전기 교체",
    "흥선교 외 3개소 보수보강공사",
}

import os, zipfile, xml.etree.ElementTree as ET
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────────────────────────────────────
# 2026년 분류 체계
# 구조: (대분류, [(중분류, [(소분류, [(부기명, [코드])]) | (부기명, [코드]) ...])])
# ─────────────────────────────────────────────────────────────────────────────

CLASSIFICATION = [
    ("법정·의무적 경비", [
        ("행정운영경비", [
            ("인력운영비", [
                ("인건비",         ["101-01","101-02","101-03","101-04"]),
                ("직무수행경비",   ["204-01","204-02"]),
                ("연금부담금",     ["304-01","304-02","304-04","307-07"]),
            ]),
            ("기본경비", [
                ("사무관리비",    ["__행정운영__201-01"]),
                ("공공운영비",    ["__행정운영__201-02"]),
                ("맞춤형복지제도",["201-04"]),
                ("국내여비",      ["__행정운영__202-01"]),
                ("업무추진비",    ["203-01","203-02","203-03","203-04"]),
            ]),
            ("의회비", [
                ("의정활동비",                    ["205-01"]),
                ("월정수당",                      ["205-02"]),
                ("의원국내여비",                   ["205-03"]),
                ("의원국외여비",                   ["205-04"]),
                ("의정운영공통경비",               ["205-05"]),
                ("의회운영업무추진비",             ["205-06"]),
                ("의원역량개발비(공공위탁·자체)",  ["205-07"]),
                ("의원역량개발비(민간위탁)",        ["205-08"]),
                ("의원정책개발비",                 ["205-09"]),
                ("의장협의체부담금",               ["205-10"]),
                ("의원국민연금부담금",             ["205-11"]),
                ("의원국민건강부담금",             ["205-12"]),
            ]),
        ]),
        ("내부거래 전출금", [
            ("기타회계·공기업특별회계 전출금", ["__701_나머지__","701-02"]),
            ("기금전출금",                     ["702-01"]),
        ]),
        ("세입관련 용도지정 세출", [
            ("국도비보조금반환금", ["802-01","802-02","802-03"]),
        ]),
        ("일반회계 예비비", [
            ("일반예비비·재난예비비·내부유보금", ["801-01","801-03"]),
        ]),
        ("지방채 원리금상환", [
            ("차입금 이자상환", ["311-02","311-03","311-05"]),
            ("차입금 원금상환", ["601-05"]),
        ]),
    ]),
    ("고정적 경비", [
        ("공단 전출금 및 출연금", [
            ("출연금",    ["306-01"]),
            ("공단전출금",["309-01"]),
        ]),
        ("법령·조례·협약사항", [
            ("사회보장적수혜금",           ["301-01","301-02","301-03"]),
            ("장학금및학자금",             ["301-04","309-02"]),
            ("자율방범대실비지원",         ["301-06"]),
            ("통장활동보상금",             ["301-07"]),
            ("민간인국외여비·외빈초청여비",["301-08","301-09"]),
            ("사회복무요원보상금",         ["301-10"]),
            ("예술단원·운동부등보상금",    ["301-12"]),
            ("기타보상금",                 ["301-14"]),
            ("이주및재해보상금",           ["302-02"]),
            ("배상금",                     ["305-01"]),
            ("자치단체간부담금",           ["__308_나머지__"]),
            ("교육기관에대한보조",         ["308-08"]),
            ("교육비특별회계전출금",       ["308-10"]),
            ("지역대학에대한경상보조",     ["308-09"]),
            ("국제부담금",                 ["310-02"]),
        ]),
        ("경직성 경상경비", [
            ("일반운영비(사무관리비·공공운영비)", ["__사업__201-01","__사업__201-02"]),
            ("행사운영비",                          ["201-03"]),
            ("여비(국내·교육)",                     ["__사업__202-01","202-05"]),
            ("국외여비",                 ["202-03","202-04"]),
            ("재료비",                   ["206-01"]),
            ("연구개발비",               ["207-01","207-02"]),
            ("포상금",                   ["303-01"]),
            ("자산취득비 및 도서구입비", ["405-01","405-02"]),
            ("기타자본이전(임차보증금)", ["406-01"]),
        ]),
        ("경직성 이전경비", [
            ("의용소방대지원",               ["301-05"]),
            ("행사실비지원금",               ["301-11"]),
            ("의료및구료비",                 ["307-01"]),
            ("지방보조금",                   ["307-02","307-03","307-04",
                                              "307-10","307-11","402-01"]),
            ("민간위탁금",                   ["__305_나머지__"]),
            ("민간위탁사업비",               ["402-03"]),
            ("보험금·민간인위탁교육비",       ["307-06","307-12"]),
            ("운수업계보조금",               ["__309_유류__"]),
            ("예비군육성지원경상보조",        ["308-12"]),
            ("공기관등에대한경상적위탁사업비",["308-13"]),
            ("공기관등에대한자본적위탁사업비",["403-02"]),
            ("기타부담금",                   ["308-14"]),
        ]),
        ("도시시설유지비", [
            # 401-01 중 노란색 셀로 구분 → 별도 추출
            ("도시기본기능유지(401-01 산출근거명 기준)", ["__도시시설유지__"]),
            # 대중교통비용: 701-01(철도교통과) + 308-07(철도/버스정책과) + 307-09(유류제외)
            ("대중교통비용",                  ["__대중교통__"]),
            # 폐기물처리비용: 307-05 중 폐기물 3개 세부사업
            ("폐기물처리비용",                ["__폐기물처리__"]),
        ]),
    ]),
    ("시설투자사업", [
        ("시설투자사업", [
            ("시설비",              ["__시설비_비노랑__"]),   # 401-01 중 노란색 제외
            ("감리비",              ["401-02"]),
            ("시설부대비",          ["401-03"]),
             # 경직성이전에도 있으나 여기선 제거

            ("공사공단자본전출금",  ["404-01"]),
        ]),
    ]),
]

# 접두어 정의
PREFIX_행정 = "__행정운영__"
PREFIX_사업  = "__사업__"
PREFIX_도시  = "__도시시설유지__"
PREFIX_시설  = "__시설비_비노랑__"
PREFIX_교통  = "__대중교통__"      # 701-01(철도교통과) + 308-07(철도/버스) + 307-09(유류제외)
PREFIX_폐기물 = "__폐기물처리__"   # 307-05 중 폐기물 3개 세부사업

# 특수코드
SPECIAL_CODES = {"201-01","201-02","202-01","401-01","701-01","308-07","307-05","307-09"}


def collect_classified_codes(cls):
    codes = set()
    for _d1, d2_list in cls:
        for _d2, items in d2_list:
            fv = items[0][1] if items else []
            has_sub = bool(fv) and isinstance(fv[0], tuple)
            if has_sub:
                for _sub, detail in items:
                    for _nm, item_codes in detail:
                        for c in item_codes:
                            if not c.startswith("__"):
                                codes.add(c)
            else:
                for _nm, item_codes in items:
                    for c in item_codes:
                        if not c.startswith("__"):
                            codes.add(c)
    codes.update(SPECIAL_CODES)
    return codes


# ─────────────────────────────────────────────────────────────────────────────
# 노란색 셀 추출 (XML 직접 파싱)
# ─────────────────────────────────────────────────────────────────────────────

def extract_urban_infra_401(xlsx_path, sheet_idx=1):
    """자체사업 시트에서 401-01 중 노란색(FFFFFF00) 산출근거명의 예산액 합계 반환"""
    YELLOW = "rgb:FFFFFF00"
    with zipfile.ZipFile(xlsx_path) as z:
        ns = {'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

        with z.open('xl/sharedStrings.xml') as f:
            ss_tree = ET.parse(f)
        shared_strings = []
        for si in ss_tree.getroot().findall('.//x:si', ns):
            text = ''.join(t.text or '' for t in si.findall('.//x:t', ns))
            shared_strings.append(text)

        with z.open('xl/styles.xml') as f:
            styles_tree = ET.parse(f)
        sroot = styles_tree.getroot()
        fills = []
        for fill in sroot.findall('.//x:fills/x:fill', ns):
            pf = fill.find('x:patternFill', ns)
            if pf is None: fills.append('none'); continue
            fg = pf.find('x:fgColor', ns)
            if fg is None: fills.append('nocolor'); continue
            rgb = fg.get('rgb'); theme = fg.get('theme'); tint = fg.get('tint')
            if rgb: fills.append(f'rgb:{rgb}')
            elif theme: fills.append(f'theme:{theme}' + (f',tint:{tint}' if tint else ''))
            else: fills.append('nocolor')
        xfs = [int(xf.get('fillId', 0)) for xf in sroot.findall('.//x:cellXfs/x:xf', ns)]

        sheet_file = f'xl/worksheets/sheet{sheet_idx+1}.xml'
        with z.open(sheet_file) as f:
            ws_tree = ET.parse(f)
        ws_root = ws_tree.getroot()

        # 워크북에서 시트 순서 확인
        # 헤더로 열 위치 확인
        stat_col = 'K'; 산출col = 'M'; amt_col = 'O'

        import re as _re
        urban_self = urban_sub = 0
        non_urban_self = non_urban_sub = 0

        for row_el in ws_root.findall('.//x:sheetData/x:row', ns):
            rn = int(row_el.get('r', 0))
            if rn < 3: continue

            def get_val(col):
                c = row_el.find(f".//x:c[@r='{col}{rn}']", ns)
                if c is None: return ''
                v = c.find('x:v', ns)
                if v is None: return ''
                return shared_strings[int(v.text)] if c.get('t')=='s' else v.text

            stat = str(get_val(stat_col)).strip()
            if stat != '401-01': continue

            amt_str = get_val(amt_col)
            amt = int(float(amt_str)) if amt_str else 0

            def get_num(col):
                v = get_val(col)
                try: return float(v)
                except: return 0.0
            is_자체 = (get_num('P') + get_num('Q') + get_num('R') + get_num('T')) == 0

            # 산출근거명 정규화: ⊙/ｏ 기호 제거 + 연속 공백 단일화
            raw_nm = str(get_val('M')).strip()
            clean_nm = _re.sub(r'\s+', ' ', raw_nm.lstrip('⊙ｏ').strip())
            is_urban = clean_nm in URBAN_INFRA_NAMES

            if is_urban:
                if is_자체: urban_self += amt
                else:        urban_sub  += amt
            else:
                if is_자체: non_urban_self += amt
                else:        non_urban_sub  += amt

    return {
        'yellow_self': urban_self,     'yellow_sub': urban_sub,
        'non_yellow_self': non_urban_self, 'non_yellow_sub': non_urban_sub,
    }


# ─────────────────────────────────────────────────────────────────────────────
# 스타일
# ─────────────────────────────────────────────────────────────────────────────

def S(s="thin", c="BBBBBB"): return Side(style=s, color=c)
def B(c="BBBBBB"): s=S(c=c); return Border(left=s, right=s, top=s, bottom=s)
def F(h): return PatternFill("solid", fgColor=h)
def Ft(color="000000", bold=False, size=9): return Font(color=color, bold=bold, size=size, name="맑은 고딕")
def Al(h="left", v="center", wrap=False, ind=0): return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=ind)

def set_row(ws, rn, col_data, bg, fg, bold=False, size=9,
            num_cols=None, col_bg=None, height=None):
    if num_cols is None: num_cols = set()
    for col, val in col_data:
        c = ws.cell(row=rn, column=col, value=val)
        cbg = (col_bg or {}).get(col, bg)
        c.font      = Ft(fg, bold, size)
        c.fill      = F(cbg)
        c.border    = B()
        c.alignment = Al("right") if col in num_cols else Al("left")
        if col in num_cols and isinstance(val, (int, float)):
            c.number_format = "#,##0"
    if height: ws.row_dimensions[rn].height = height


# ─────────────────────────────────────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────────────────────────────────────

def main(input_file=None, output_file=None, year=None):
    # GUI 또는 CLI에서 파라미터를 받거나, 없으면 기본값 사용
    INPUT  = input_file  or "/mnt/user-data/uploads/합본예산서최종_세출__20261추__편집완료.xlsx"
    OUTPUT = output_file or "/mnt/user-data/outputs/세출_성질별_분류_2026.xlsx"
    os.makedirs(os.path.dirname(os.path.abspath(OUTPUT)), exist_ok=True)
    YEAR = year or 2026

    # ── 1. 워크북 시트 순서 확인 ──────────────────────────────────────────────
    print("[1/5] 시트 구조 확인...")
    with zipfile.ZipFile(INPUT) as z:
        ns = {'x': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        with z.open('xl/workbook.xml') as f:
            wb_tree = ET.parse(f)
        sheets = [(sh.get('name'), int(sh.get('sheetId')))
                  for sh in wb_tree.getroot().findall('.//x:sheet', ns)]
    print(f"  시트 목록: {sheets}")

    # 자체사업 시트 인덱스 찾기 (0-based)
    자체_idx = next((i for i, (nm, _) in enumerate(sheets) if '자체' in nm), 1)
    print(f"  자체사업 시트: idx={자체_idx} ({sheets[자체_idx][0]})")

    # ── 2. 데이터 로드 (일반회계 자체사업 시트) ───────────────────────────────
    print("[2/5] 데이터 로드 중...")
    sheet_name = sheets[자체_idx][0]
    df = pd.read_excel(INPUT, sheet_name=sheet_name, header=1,
                       dtype={"통계목": str})
    for c in ["최종예산액","국고보조금","균특보조금","기금보조금",
              "특별교부세","광역보조금","특별조정금","자체재원","지방채","기타"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    df["통계목"] = df["통계목"].astype(str).str.strip()
    보조컬럼 = ["국고보조금","균특보조금","기금보조금","광역보조금"]
    기존보조 = [c for c in 보조컬럼 if c in df.columns]
    df["is_자체"] = (df[기존보조].sum(axis=1) == 0) if 기존보조 else True
    df["is_행정운영"] = df["정책사업명"].astype(str).str.contains("행정운영경비", na=False) if "정책사업명" in df.columns else False

    amt_col = "최종예산액"
    G_ALL  = int(df[amt_col].sum())
    G_SELF = int(df[df["is_자체"]][amt_col].sum())
    G_SUB  = int(df[~df["is_자체"]][amt_col].sum())
    print(f"  합계: {G_ALL:,}  자체: {G_SELF:,}  보조: {G_SUB:,}")

    # ── 3. 통계목별 집계 ──────────────────────────────────────────────────────
    print("[3/5] 통계목별 집계 중...")

    def agg(mask):
        g = df[mask].groupby(["통계목","is_자체"])[amt_col].sum().unstack(fill_value=0)
        for b in [True, False]:
            if b not in g.columns: g[b] = 0
        return g[True].to_dict(), g[False].to_dict()

    STAT_SELF, STAT_SUB   = agg(pd.Series([True]*len(df), index=df.index))
    STAT_행정SELF, STAT_행정SUB = agg(df["is_행정운영"])
    STAT_사업SELF, STAT_사업SUB = agg(~df["is_행정운영"])
    STAT_NM = df.drop_duplicates("통계목").set_index("통계목")["통계목명"].to_dict() if "통계목명" in df.columns else {}

    # ── 특수 필터 집계 ──────────────────────────────────────────────────────
    def agg_filter(mask):
        sub = df[mask].copy()
        g = sub.groupby("is_자체")[amt_col].sum()
        return int(g.get(True, 0)), int(g.get(False, 0))

    # (1)+(3) 대중교통비용
    mask_교통_701 = (df["통계목"] == "701-01") & (df["부서명"].astype(str) == "철도교통과")
    mask_교통_308 = (df["통계목"] == "308-07") & (df["부서명"].astype(str).isin(["철도교통과","버스정책과"]))
    mask_교통_309 = (df["통계목"] == "307-09") & (~df["산출근거명"].astype(str).str.contains("유류세 연동보조금", na=False))
    mask_교통 = mask_교통_701 | mask_교통_308 | mask_교통_309
    교통_SELF, 교통_SUB = agg_filter(mask_교통)

    # (4) 폐기물처리비용
    폐기물_사업명 = ["폐기물 수거", "환경자원센터 운영", "폐기물처리시설 운영"]
    mask_폐기물 = (df["통계목"] == "307-05") & (df["세부사업명"].astype(str).isin(폐기물_사업명))
    폐기물_SELF, 폐기물_SUB = agg_filter(mask_폐기물)

    # 701-01: 철도교통과 제외분 (나머지는 기존 내부거래 전출금)
    mask_701_나머지 = (df["통계목"] == "701-01") & (~(df["부서명"].astype(str) == "철도교통과"))
    v701_나머지_SELF, v701_나머지_SUB = agg_filter(mask_701_나머지)

    # 308-07: 교통과 제외분 (나머지는 기존 자치단체간부담금)
    mask_308_나머지 = (df["통계목"] == "308-07") & (~df["부서명"].astype(str).isin(["철도교통과","버스정책과"]))
    v308_나머지_SELF, v308_나머지_SUB = agg_filter(mask_308_나머지)

    # 307-05: 폐기물 제외분 (나머지는 기존 민간위탁금)
    mask_305_나머지 = (df["통계목"] == "307-05") & (~df["세부사업명"].astype(str).isin(폐기물_사업명))
    v305_나머지_SELF, v305_나머지_SUB = agg_filter(mask_305_나머지)

    # 307-09: 유류세분 (기존 운수업계보조금으로 남길 것)
    mask_309_유류 = (df["통계목"] == "307-09") & (df["산출근거명"].astype(str).str.contains("유류세 연동보조금", na=False))
    v309_유류_SELF, v309_유류_SUB = agg_filter(mask_309_유류)

    print(f"  대중교통비용: 자체={교통_SELF:,} / 보조={교통_SUB:,}")
    print(f"  폐기물처리비용: 자체={폐기물_SELF:,} / 보조={폐기물_SUB:,}")

    # ── 4. 노란색 401-01 추출 ────────────────────────────────────────────────
    print("[4/5] 노란색 셀(도시시설유지비) 추출 중...")
    try:
        yellow_info = extract_urban_infra_401(INPUT, sheet_idx=자체_idx)
        YELLOW_SELF = yellow_info['yellow_self']
        YELLOW_SUB  = yellow_info['yellow_sub']
        NON_YELLOW_SELF = yellow_info['non_yellow_self']
        NON_YELLOW_SUB  = yellow_info['non_yellow_sub']
        print(f"  도시기본기능유지(산출근거명매칭): 자체={YELLOW_SELF:,} / 보조={YELLOW_SUB:,}")
        print(f"  시설투자(비노란색):     자체={NON_YELLOW_SELF:,} / 보조={NON_YELLOW_SUB:,}")
    except Exception as e:
        print(f"  노란색 추출 실패({e}), 401-01 전체를 시설투자로 처리")
        YELLOW_SELF = YELLOW_SUB = 0
        NON_YELLOW_SELF = int(STAT_SELF.get('401-01', 0))
        NON_YELLOW_SUB  = int(STAT_SUB.get('401-01', 0))

    classified_codes = collect_classified_codes(CLASSIFICATION)
    all_codes = set(df["통계목"].unique()) - {"nan","None",""}
    unclassified_codes = sorted(all_codes - classified_codes)

    # ── 5. 집계 함수 ──────────────────────────────────────────────────────────
    # 특수 필터 맵 (접두어 → (self, sub))
    SPECIAL_MAP = {
        PREFIX_도시:   (YELLOW_SELF,    YELLOW_SUB),
        PREFIX_시설:   (NON_YELLOW_SELF, NON_YELLOW_SUB),
        PREFIX_교통:   (교통_SELF,      교통_SUB),
        PREFIX_폐기물: (폐기물_SELF,    폐기물_SUB),
        # 나머지 분 (기존 항목에서 특수분 차감)
        "__701_나머지__": (v701_나머지_SELF, v701_나머지_SUB),
        "__308_나머지__": (v308_나머지_SELF, v308_나머지_SUB),
        "__305_나머지__": (v305_나머지_SELF, v305_나머지_SUB),
        "__309_유류__":   (v309_유류_SELF,  v309_유류_SUB),
    }

    def get_vs_vb(codes):
        vs = vb = 0
        for c in codes:
            if c in SPECIAL_MAP:
                sv, sb = SPECIAL_MAP[c]
                vs += sv; vb += sb
            elif c.startswith(PREFIX_행정):
                real = c[len(PREFIX_행정):]
                vs += int(STAT_행정SELF.get(real, 0))
                vb += int(STAT_행정SUB.get(real, 0))
            elif c.startswith(PREFIX_사업):
                real = c[len(PREFIX_사업):]
                vs += int(STAT_사업SELF.get(real, 0))
                vb += int(STAT_사업SUB.get(real, 0))
            else:
                vs += int(STAT_SELF.get(c, 0))
                vb += int(STAT_SUB.get(c, 0))
        return vs, vb

    def disp_codes(codes):
        if not codes: return "통계목 미기재"
        DISP_MAP = {
            PREFIX_도시:        "401-01(노란색)",
            PREFIX_시설:        "401-01(노란색 제외)",
            PREFIX_교통:        "701-01(철도교통과)+308-07(철도/버스정책과)+307-09(유류제외)",
            PREFIX_폐기물:      "307-05(폐기물수거·환경자원센터·폐기물처리시설)",
            "__701_나머지__":   "701-01(철도교통과 제외)",
            "__308_나머지__":   "308-07(철도·버스정책과 제외)",
            "__305_나머지__":   "307-05(폐기물 제외)",
            "__309_유류__":     "307-09(유류세연동보조금만)",
        }
        out = []
        for c in codes:
            if c in DISP_MAP: out.append(DISP_MAP[c])
            elif c.startswith(PREFIX_행정): out.append(c[len(PREFIX_행정):] + "(행정운영경비內)")
            elif c.startswith(PREFIX_사업): out.append(c[len(PREFIX_사업):] + "(사업비內)")
            else: out.append(c)
        return ", ".join(out)

    # ── 6. 엑셀 작성 ──────────────────────────────────────────────────────────
    print("[5/5] 엑셀 작성 중...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "성질별_분류"
    NUM = {3, 4, 5}

    # 제목
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value     = f"{YEAR}년도 세출예산 성질별 분류 (일반회계 자체사업 기준)"
    c.font      = Ft("FFFFFF", bold=True, size=13)
    c.fill      = F("1F3864")
    c.alignment = Al("center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value = ("※ 자체사업: 국고+균특+기금+광역보조금 합계=0  |  "
               "도시시설유지비: 401-01 중 노란색(FFFFFF00) 셀  |  단위: 천원")
    c.font      = Ft("555555", size=8)
    c.alignment = Al("right")
    ws.row_dimensions[2].height = 13

    for col, h in enumerate(["구분","해당 통계목","자체사업","보조사업","합계"], 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font      = Ft("FFFFFF", bold=True, size=10)
        c.fill      = F("1F3864")
        c.border    = B()
        c.alignment = Al("center")
    ws.row_dimensions[3].height = 20

    row = 4
    tot_cls_self = tot_cls_sub = 0

    for d1_name, d2_list in CLASSIFICATION:
        d1_r = row; row += 1
        d1_s = d1_b = 0

        for d2_name, items in d2_list:
            d2_r = row; row += 1
            d2_s = d2_b = 0

            fv = items[0][1] if items else []
            has_sub = bool(fv) and isinstance(fv[0], tuple)

            if has_sub:
                for sub_name, detail in items:
                    s3_r = row; row += 1
                    s3_s = s3_b = 0
                    for item_name, codes in detail:
                        vs, vb = get_vs_vb(codes)
                        s3_s += vs; s3_b += vb
                        set_row(ws, row,
                                [(1,f"      {item_name}"), (2, disp_codes(codes)),
                                 (3,vs),(4,vb),(5,vs+vb)],
                                bg="FFFFFF", fg="333333", size=9, num_cols=NUM,
                                col_bg={2:"F7F7F7",3:"EBF5FB",4:"FEF9E7",5:"FDEBD0"}, height=15)
                        row += 1
                    d2_s += s3_s; d2_b += s3_b
                    set_row(ws, s3_r,
                            [(1,f"    {sub_name}"),(2,""),
                             (3,s3_s),(4,s3_b),(5,s3_s+s3_b)],
                            bg="BDD7EE", fg="1A3A5C", bold=True, size=9, num_cols=NUM,
                            col_bg={3:"9FC5E8",4:"8DB8D9",5:"7AAEC9"}, height=17)
            else:
                for item_name, codes in items:
                    vs, vb = get_vs_vb(codes)
                    d2_s += vs; d2_b += vb
                    set_row(ws, row,
                            [(1,f"    {item_name}"), (2, disp_codes(codes)),
                             (3,vs),(4,vb),(5,vs+vb)],
                            bg="FFFFFF", fg="333333", size=9, num_cols=NUM,
                            col_bg={2:"F7F7F7",3:"EBF5FB",4:"FEF9E7",5:"FDEBD0"}, height=15)
                    row += 1

            d1_s += d2_s; d1_b += d2_b
            set_row(ws, d2_r,
                    [(1,f"  {d2_name}"),(2,""),
                     (3,d2_s),(4,d2_b),(5,d2_s+d2_b)],
                    bg="2E75B6", fg="FFFFFF", bold=True, size=9, num_cols=NUM,
                    col_bg={3:"1A6090",4:"154E75",5:"0F3A5C"}, height=18)

        tot_cls_self += d1_s; tot_cls_sub += d1_b
        set_row(ws, d1_r,
                [(1,d1_name),(2,""),
                 (3,d1_s),(4,d1_b),(5,d1_s+d1_b)],
                bg="1F3864", fg="FFFFFF", bold=True, size=10, num_cols=NUM,
                col_bg={3:"163257",4:"0F2440",5:"09192E"}, height=21)

    # 미분류 섹션
    unc_d1_r = row; row += 1
    unc_s = unc_b = 0
    for code in unclassified_codes:
        vs = int(STAT_SELF.get(code, 0)); vb = int(STAT_SUB.get(code, 0))
        unc_s += vs; unc_b += vb
        nm = STAT_NM.get(code, "")
        set_row(ws, row,
                [(1,f"    {code}  {nm}"),(2,code),(3,vs),(4,vb),(5,vs+vb)],
                bg="FFF8E1", fg="7D4B00", size=9, num_cols=NUM,
                col_bg={2:"FFF3CD",3:"FDEBD0",4:"FAE5D3",5:"F9E4B7"}, height=15)
        row += 1
    set_row(ws, unc_d1_r,
            [(1,"★ 미분류 통계목 (분류 체계 미포함)"),(2,""),
             (3,unc_s),(4,unc_b),(5,unc_s+unc_b)],
            bg="D68910", fg="FFFFFF", bold=True, size=10, num_cols=NUM,
            col_bg={3:"B7770D",4:"9C640A",5:"7D5108"}, height=21)

    # 검증
    row += 1
    final_self = tot_cls_self + unc_s
    final_sub  = tot_cls_sub  + unc_b
    diff_s = G_SELF - final_self; diff_b = G_SUB - final_sub
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1,
                value=(f"[집계 검증]  원본합계: {G_ALL:,}  |  "
                       f"분류합계: {final_self+final_sub:,}  |  "
                       f"차이(자체): {diff_s:,} / 차이(보조): {diff_b:,}  "
                       + ("✔ 일치" if diff_s==0 and diff_b==0 else "✘ 불일치")))
    c.font      = Ft("1A3A5C", bold=True, size=9)
    c.fill      = F("D6EAF8" if diff_s==0 else "FADBD8")
    c.border    = B("2E75B6")
    c.alignment = Al("center")
    ws.row_dimensions[row].height = 16
    row += 1

    # 총합계
    set_row(ws, row,
            [(1,"▶ 총합계 (분류 + 미분류)"),(2,""),
             (3,final_self),(4,final_sub),(5,final_self+final_sub)],
            bg="C0392B", fg="FFFFFF", bold=True, size=10, num_cols=NUM,
            col_bg={3:"A93226",4:"922B21",5:"7B241C"}, height=24)

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.freeze_panes = "A4"

    # 시트2: 통계목별 원시집계
    ws2 = wb.create_sheet("통계목별_원시집계")
    _raw_sheet(ws2, STAT_SELF, STAT_SUB, STAT_NM, classified_codes, unclassified_codes, YEAR, amt_col)

    wb.save(OUTPUT)
    print(f"\n저장 완료 → {OUTPUT}")
    print()
    W = 24
    print(f"  {'구분':<{W}} {'자체사업':>18} {'보조사업':>18} {'합계':>18}")
    print(f"  {'─'*80}")
    print(f"  {'분류 체계 집계':<{W}} {tot_cls_self:>18,} {tot_cls_sub:>18,} {tot_cls_self+tot_cls_sub:>18,}")
    print(f"  {'미분류 집계':<{W}} {unc_s:>18,} {unc_b:>18,} {unc_s+unc_b:>18,}")
    print(f"  {'─'*80}")
    print(f"  {'최종 합계':<{W}} {final_self:>18,} {final_sub:>18,} {final_self+final_sub:>18,}")
    print(f"  {'원본 합계':<{W}} {G_SELF:>18,} {G_SUB:>18,} {G_ALL:>18,}")
    print(f"  {'차이':<{W}} {diff_s:>18,} {diff_b:>18,}")
    if diff_s == 0 and diff_b == 0:
        print("\n  ✔ 집계 일치 확인 완료")
    else:
        print("\n  ✘ 집계 불일치 — 미분류 통계목 확인 필요")


def _raw_sheet(ws, stat_self, stat_sub, stat_nm, classified_codes, unclassified_codes, year, amt_col):
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = f"{year}년 통계목별 집계 (단위: 천원)"
    c.font      = Ft("FFFFFF", bold=True, size=12)
    c.fill      = F("1F3864")
    c.alignment = Al("center")
    ws.row_dimensions[1].height = 24
    for col, h in enumerate(["통계목","통계목명","자체사업","보조사업","합계","분류여부"], 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Ft("FFFFFF", bold=True); c.fill = F("2E75B6")
        c.border = B(); c.alignment = Al("center")
    ws.row_dimensions[2].height = 18
    all_codes = sorted(set(list(stat_self.keys()) + list(stat_sub.keys())))
    ts = tb = 0
    for i, code in enumerate(all_codes, 3):
        vs = int(stat_self.get(code, 0)); vb = int(stat_sub.get(code, 0))
        ts += vs; tb += vb
        nm  = stat_nm.get(code, "")
        unc = code in unclassified_codes
        bg  = "FFF8E1" if unc else ("F0F8FF" if i%2==0 else "FFFFFF")
        lbl = "미분류" if unc else "분류됨"
        for col, val in [(1,code),(2,nm),(3,vs),(4,vb),(5,vs+vb),(6,lbl)]:
            c = ws.cell(row=i, column=col, value=val)
            c.fill = F(bg); c.border = B()
            c.alignment = Al("right") if col in {3,4,5} else Al("left")
            if col in {3,4,5}: c.number_format = "#,##0"
    r = len(all_codes) + 3
    for col, val in [(1,"합계"),(2,""),(3,ts),(4,tb),(5,ts+tb),(6,"")]:
        c = ws.cell(row=r, column=col, value=val)
        c.font = Ft("FFFFFF", bold=True); c.fill = F("1F3864")
        c.border = B()
        c.alignment = Al("right") if col in {3,4,5} else Al("left")
        if col in {3,4,5}: c.number_format = "#,##0"
    for col, w in [(1,12),(2,30),(3,18),(4,18),(5,18),(6,10)]:
        ws.column_dimensions[chr(64+col)].width = w
    ws.freeze_panes = "A3"


if __name__ == "__main__":
    main()
